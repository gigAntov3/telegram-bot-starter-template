import openpyxl
import openpyxl.cell
from datetime import date
import asyncio

from bot.__init__ import db

class UsersExcel:

    def __init__(self) -> None:
        self.wb = openpyxl.Workbook()
        self.user_sheet = self.wb.create_sheet("Пользователи", 0)
        self.wb.remove(self.wb["Sheet"])
        

    def marking_user(self):
        self.user_sheet.cell(row=1, column=1, value="ID")
        self.user_sheet.cell(row=1, column=2, value="Telegram ID")
        self.user_sheet.cell(row=1, column=3, value="Имя")
        self.user_sheet.cell(row=1, column=4, value="Ссылка")
        self.user_sheet.cell(row=1, column=6, value="Дата регистрации")
    
    async def copy_user(self):
        count = 0
        row = 2
        while(True):
            datas= await db.get_users(count, 50)
            if(datas == []): break
            for data in datas:
                self.user_sheet.cell(row=row, column=1, value=data[0])
                self.user_sheet.cell(row=row, column=2, value=data[1])
                self.user_sheet.cell(row=row, column=3, value=data[2])
                self.user_sheet.cell(row=row, column=4, value=data[3])
                self.user_sheet.cell(row=row, column=6, value=data[4].strftime("%d.%m.%Y"))
                row += 1
            count += 50

        
    async def create(self):
        self.marking_user()
        await self.copy_user()
        self.wb.save("data/statistics.xlsx")


if __name__ == "__main__":    
    ex = UsersExcel()
    asyncio.run(ex.create())